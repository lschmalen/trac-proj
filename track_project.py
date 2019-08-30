import wx
import wx.lib.buttons
import datetime
import configparser
import math
import glob
import os
import shutil
import re
import csv
from openpyxl import Workbook
from functools import reduce
from collections import Counter
 
 
def Warn(parent, message, caption = 'Warning!'):
	dlg = wx.MessageDialog(parent, message, caption, wx.OK | wx.ICON_WARNING)
	dlg.ShowModal()
	dlg.Destroy()


class SettingsDialog(wx.Dialog):
	""" This window displays the settings that can be used to configure the buttons """
	def __init__(self, *args, **kwargs):
		wx.Dialog.__init__(self, *args, **kwargs)

		vbox = wx.BoxSizer(wx.VERTICAL)
		tt = wx.StaticText(self, -1, "Buttons") 		
		vbox.Add(tt, -1, wx.EXPAND | wx.ALIGN_LEFT | wx.ALL, 5)

		self.ti = wx.TextCtrl(self, size = (200,120),style = wx.TE_MULTILINE) 		
		vbox.Add(self.ti, 1 ,wx.EXPAND | wx.ALIGN_LEFT | wx.ALL, 0)

		hbox = wx.BoxSizer(wx.HORIZONTAL)
		saveButton = wx.Button(self, label='Save && Close')
		saveButton.Bind(wx.EVT_BUTTON, self.OnSave)
		cancelButton = wx.Button(self, label='Cancel')
		cancelButton.Bind(wx.EVT_BUTTON, self.OnCancel)
		hbox.Add(saveButton, -1, wx.ALL, 5)
		hbox.Add(cancelButton, -1, wx.ALL, 5)

		vbox.Add(hbox, 1, wx.EXPAND | wx.ALL, 5)
		self.SetSizer(vbox)

		# Fill in text control field with entries from ini file
		config = configparser.ConfigParser()
		config.optionxform = str 
		config.read('config.ini')
		if config.has_section('Buttons') == True:		
			button_dict = dict(config['Buttons'])
			for ii in range(len(button_dict)):
				self.ti.AppendText('%s\n' % button_dict['Button%d' % (ii+1)])
			
	def OnCancel(self, event):
		self.Destroy()

	def OnSave(self, event):
		# save ini file
		config = configparser.ConfigParser()
		config.optionxform = str 
		config.read('config.ini')

		new_button_dict = dict()
		count = 1
		for ii in range(self.ti.GetNumberOfLines()):
			if self.ti.GetLineLength(ii) > 0:
				new_button_dict['Button%d' % count] = self.ti.GetLineText(ii)
				count += 1
		if count > 21:
			Warn(self,'Too many categories (max 20)')
		else:
			config['Buttons'] = new_button_dict
			with open('config.ini', 'w') as config_file:
				config.write(config_file)		
		
			Warn(self, 'Restart program for changes to become effective')
			self.Destroy()



# Main Window
class ProjectFrame(wx.Frame):
	""" This window displays a set of buttons curresponding to the different projects """
	def __init__(self, *args, **kwargs):
		wx.Frame.__init__(self, *args, **kwargs)
			
		self.read_config_file()
		if len(self.button_name_list) == 0:
			self.Close()
		
		print(self.color_list)
		if len(self.button_name_list) > 20:
			Warn(self, 'Too many categories (max 20)')
			self.Close()

		# check if datafiles directory exists, otherwise create it
		try:
			os.makedirs("datafiles")	
		except FileExistsError:
			# directory exists
			pass
		
		
		# generate list with buttons
		button_height = round(800 / min(10,len(self.button_name_list)))
		button_width = round(3.5 * button_height)
		
		self.old_active_button = -1
		self.start_time = -1
		self.btn_list = [wx.lib.buttons.GenToggleButton(self,label=bn,size=(button_width,button_height)) for bn in self.button_name_list]
		
		# grid sizer		
		if len(self.button_name_list) <= 10:			
			sizer_toggle = [wx.BoxSizer(wx.VERTICAL)] 
			first_rows = len(self.button_name_list)
		else:
			sizer_toggle = [wx.BoxSizer(wx.VERTICAL), wx.BoxSizer(wx.VERTICAL)]
			first_rows = math.ceil(len(self.button_name_list)/2.0)
		
		# add to respective sizers
		for i in range(len(self.button_name_list)):			
			self.btn_list[i].Bind(wx.EVT_BUTTON, lambda evt, ai=i: self.OnButton(evt, ai) )
			self.btn_list[i].SetBackgroundColour(self.color_list[i])
			self.btn_list[i].SetForegroundColour('#000000')
			self.btn_list[i].SetFont(wx.Font(22, wx.SWISS, wx.NORMAL, wx.NORMAL))
			self.btn_list[i].SetValue(False)			
			sizer_toggle[i//first_rows].Add(self.btn_list[i], 0, wx.ALL, 5)
		
			
		# add extra buttons
		sizer_extra = wx.BoxSizer(wx.VERTICAL)
		
		# Self-explanatory, exists the program
		exit_button = wx.Button(self,label='Exit',size=(button_width//2,button_height//2))
		exit_button.Bind(wx.EVT_BUTTON, self.OnExit)
		exit_button.SetFont(wx.Font(16, wx.SWISS, wx.NORMAL, wx.NORMAL))
		sizer_extra.Add(exit_button, 0, wx.ALL, 10)
		
		# Converts the datebase into an excel-sheet
		convert_button = wx.Button(self, label='Convert', size=(button_width//2,button_height//2))
		convert_button.Bind(wx.EVT_BUTTON, self.OnConvert)
		convert_button.SetFont(wx.Font(16, wx.SWISS, wx.NORMAL, wx.NORMAL))
		sizer_extra.Add(convert_button, 0, wx.ALL, 10)
		
		# consolidate the database by grouping identical entries that happened on the same day. Will destroy the sequence, so use with care
		consolidate_button = wx.Button(self, label='Consolidate', size=(button_width//2,button_height//2))
		consolidate_button.Bind(wx.EVT_BUTTON, self.OnConsolidate)
		consolidate_button.SetFont(wx.Font(16, wx.SWISS, wx.NORMAL, wx.NORMAL))
		sizer_extra.Add(consolidate_button, 0, wx.ALL, 10)		
		
		# settings button
		settings_button = wx.Button(self, label='Settings')
		settings_button.Bind(wx.EVT_BUTTON, self.OnSettings)
		sizer_extra.Add(settings_button, 0, wx.ALL | wx.ALIGN_BOTTOM | wx.EXPAND, 10)

		# put everything together
		sizer_global = wx.BoxSizer(wx.HORIZONTAL)
		for i in sizer_toggle: sizer_global.Add(i, 0, wx.ALL, 0)
		sizer_global.Add(sizer_extra, 0, wx.ALL, 20)
		self.SetSizerAndFit(sizer_global)
		


		
	# the the config.ini file and generate the buttons
	def read_config_file(self):
		self.button_name_list = []
		# colors are the tab20 colormap from matplotlib
		self.color_list = ['#1f77b4','#ff7f0e','#2ca02c','#d62728','#9467bd','#8c564b','#e377c2','#7f7f7f','#bcbd22','#17becf','#aec7e8','#ffbb78','#98df8a','#ff9896','#c5b0d5','#c49c94','#f7b6d2','#c7c7c7','#dbdb8d','#9edae5']	
		
		config = configparser.ConfigParser()
		config.optionxform = str 
		config.read('config.ini')
		if config.has_section('Buttons') == False:
			print('Buttons section not found in INI')
			Warn(self, 'Buttons section not found in INI')
			return retval,color_list
				
		Button_dict = (dict(config['Buttons']))
		for i in range(len(self.color_list)):
			bs = 'Button%d'%(i+1)
			if bs in Button_dict: self.button_name_list.append(config['Buttons'][bs])
			cs = 'Color%d'%(i+1)
			if cs in Button_dict: self.color_list[i] = config['Buttons'][cs]
		if 'Button%d' % (len(self.color_list)+1) in Button_dict:
			Warn(self, 'Only %d buttons displayed, please wait for future release' % len(self.color_list))
		
		self.output_percentage = False
		if config.has_section('Output') == True:
			if 'Percentage' in config['Output']:
				self.output_percentage = config['Output'].getboolean('Percentage')
		
	# convert to percentages and make sure that sum equals 100
	def get_percentage(self, x):
		xs = [100.0 * xe / sum(x) for xe in x]
		xsr = [math.floor(xe) for xe in xs]
		while sum(xsr) < 100:
			# find smallest relative errors			
			relerr = [(x+1.0)/y if y > 0 else 9000 for x,y in zip(xsr,xs)]			
			_,idx = min( (relerr[i],i) for i in range(len(relerr)) )
			xsr[idx] += 1			
		return xsr
	
	# generate file name with the current date
	def get_file_name(self):
		today = datetime.datetime.now()
		filename = 'datafiles/track%d_%d_%d.lstrac' % (today.year, today.month, today.day)
		return filename
		
	# When a toggle button is activated	
	def OnButton(self, Event, active_index):					
		deactiveate_action = False
		for btn_i in range(len(self.btn_list)):			
			# find active button and change color if necessary			
			if btn_i == active_index:							
				# if already active, then deactivate
				if btn_i == self.old_active_button:						
					self.btn_list[btn_i].SetBackgroundColour(self.color_list[btn_i])					
					self.btn_list[btn_i].SetForegroundColour("#000000")
					self.btn_list[btn_i].SetValue(False)						
					deactiveate_action = True
				else:										
					self.btn_list[btn_i].SetBackgroundColour("#FFFFFF")
					self.btn_list[btn_i].SetForegroundColour("#FF0000")
					self.btn_list[btn_i].SetValue(True)
			else:					
				if self.btn_list[btn_i].GetValue() == True:					
					self.btn_list[btn_i].SetBackgroundColour(self.color_list[btn_i])				
					self.btn_list[btn_i].SetForegroundColour("#000000")
					self.btn_list[btn_i].SetValue(False)			
					
		
		# here the real action begins that counts the time		
		if self.old_active_button == -1:
			self.start_time = datetime.datetime.now()
		else:
			stop_time = datetime.datetime.now()
			print('Spent %d seconds on %s' % ((stop_time-self.start_time).total_seconds(), self.button_name_list[self.old_active_button]))
			# write to file						
			with open(self.get_file_name(), 'a') as fd:
				fd.write('%s,%d\n' % (self.button_name_list[self.old_active_button],(stop_time-self.start_time).total_seconds()))
				fd.close()
			# get new start time for next call
			self.start_time = datetime.datetime.now()
		
		self.old_active_button = active_index
		if deactiveate_action == True: self.old_active_button = -1

	def OnSettings(self, Event):
		settingsDialog = SettingsDialog(None, title='Settings')
		settingsDialog.ShowModal()
		settingsDialog.Destroy()
			

	def OnExit(self, Event):
		# if button is active, write it out
		if self.old_active_button >= 0:
			# write the current active one to the file
			stop_time = datetime.datetime.now()
			#print('Spent %d seconds on %s' % ((stop_time-self.start_time).total_seconds(), self.button_name_list[self.old_active_button]))
			# write to file						
			fd = open(self.get_file_name(), 'a')
			fd.write('%s,%d\n' % (self.button_name_list[self.old_active_button],(stop_time-self.start_time).total_seconds()))
			fd.close()						
		self.Close()

	# Convert the data to an Excel file with weekly and daily statistics
	def OnConvert(self, Event):
		tracking = []
		file_list = glob.glob('./datafiles/*.lstrac')
		for filename in file_list:
			match = re.search('track(\d+)_(\d+)_(\d+)',filename)
			if match:
				datum = datetime.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
				mydict = {}
				# read file into dictionary				
				reader = csv.reader(open(filename,mode='r'))
				for row in reader: 
					if not ''.join(row).strip(): continue
					if not row[0] in mydict: mydict[row[0]] = int(row[1])
					else: mydict[row[0]] += int(row[1])
				tracking.append((datum,mydict))
		
					
		years = list(set([int(a[0].year) for a in tracking]))
		wb = Workbook()
		for year in years:
			# extract tracking of given year only, sort according to ascending date
			tracking_year = sorted([t for t in tracking if int(t[0].year) == year], key=lambda x: x[0])
						
			# Build union of all used project entries	
			allkey = list(reduce(set.union, (set(d[1].keys()) for d in tracking_year)))
			
			# sort the strings by amount of seconds spent			
			c = Counter()
			for d in tracking_year: c.update(d[1])			
			sorted_projects = [x[0] for x in sorted(c.items(), key=lambda x: x[1],reverse=True)]
						
			# get new list containing the sum of 
			ws = wb.create_sheet('Daily %d' % year)
			ws.append(['Minutes spent daily in projects'])
			ws.append(['Date','KW'] + sorted_projects)			
						
			for j in range(len(tracking_year)):					
				cell = ws.cell(row=j+3, column=1, value = tracking_year[j][0])				
				cell.number_format = 'DD.MM.YYYY'
				_ = ws.cell(row=j+3, column=2, value = tracking_year[j][0].isocalendar()[1])
				for idx,project in enumerate(sorted_projects):					
					if project in tracking_year[j][1]:
						_ = ws.cell(row=j+3, column=idx+3, value = int(round(tracking_year[j][1][project] / 60.0)))
			
			# create weekly stats
			tracking_KW = [(t[0].isocalendar()[1], t[1]) for t in tracking_year]			
			unique_weeks = sorted(list(set([d[0] for d in tracking_KW])))
			ws = wb.create_sheet('Weekly %d' % year)
			if self.output_percentage == True:
				ws.append(['Percentage of time spent weekly in different projects'])
			else:
				ws.append(['Minutes spent weekly in different projects'])
			ws.append(['KW'] + sorted_projects)
			for j in range(len(unique_weeks)):
				weekly = [t for t in tracking_KW if t[0] == unique_weeks[j]]				
				cw = Counter()
				for wi in weekly: cw.update(wi[1])
				_ = ws.cell(row=j+3, column=1, value = unique_weeks[j])
				
				
				if self.output_percentage == False:
					for idx,project in enumerate(sorted_projects):
						if project in cw:
							_ = ws.cell(row=j+3,column=idx+2, value = int(round(cw[project] / 60.0)))
				else:
					sorted_minutes = [cw[project] for project in sorted_projects]
					if sum(sorted_minutes) > 0:
						sorted_percentage = self.get_percentage(sorted_minutes)
						for idx in range(len(sorted_projects)):
							_ = ws.cell(row=j+3, column = idx+2, value = sorted_percentage[idx])
			
		# Save file		
		with wx.FileDialog(self, "Save Excel file", wildcard="Excel files (*.xlsx)|*.xlsx", style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
			if fileDialog.ShowModal() == wx.ID_CANCEL:
				return     # the user changed their mind

			# save the current contents in the file
			pathname = fileDialog.GetPath()
			wb.save(pathname)
		


	# Consolidate files to save space, is not really necessary unless the project grows extremely large
	def OnConsolidate(self, Event):
		Warn(self, 'Consolidation temporarily disabled')
		return

		dlg = wx.MessageDialog(None, 'Do you really want to consolidate the database? Loss of data may occur','Consolidation',wx.YES_NO | wx.ICON_QUESTION)
		result = dlg.ShowModal()
		if result == wx.ID_YES:
			# generate backup
			today = datetime.datetime.now()
			backupdirname = 'backup_%d_%d_%d' % (today.year, today.month, today.day)
		
			try:
				os.makedirs(backupdirname)	
			except FileExistsError:
				# directory exists
				pass

			backup_list = glob.glob('./%s/*.lstrac' % backupdirname)
			for filename in backup_list:
				os.unlink(filename)
			file_list = glob.glob('./datafiles/*.lstrac')
			for filename in file_list:
				shutil.copy(filename, backupdirname)
			
			for filename in file_list:
				c = Counter()
				with open(filename,mode='r') as file:
					reader = csv.reader(file)
					for row in reader: 				
						if not ''.join(row).strip(): continue
						c.update({row[0] : int(row[1])})
				
					file.close()
				with open(filename,mode='w') as file:
					for ci in c:
						file.write('%s,%d\n'%(ci,c[ci]))
					file.close();
		return
		
app = wx.App(False)
frame = ProjectFrame(None, title="Project Time Tracker Tool")
frame.Show()

app.MainLoop()